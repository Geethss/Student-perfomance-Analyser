"""
Report Generation Module
Generates Excel and PDF reports from analysis results
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import io
from typing import List, Dict, Any
from datetime import datetime
import unicodedata
import re


class ReportGenerator:
    """Generate formatted reports in Excel and PDF formats"""
    
    @staticmethod
    def sanitize_for_pdf(text: str) -> str:
        """
        Sanitize text for PDF output by replacing/removing Unicode characters
        that aren't supported by standard PDF fonts
        
        Args:
            text: Input text that may contain Unicode characters
            
        Returns:
            Sanitized text safe for PDF rendering
        """
        # Common mathematical symbol replacements
        replacements = {
            '∫': 'integral',
            '∑': 'sum',
            '∏': 'product',
            '√': 'sqrt',
            '∞': 'infinity',
            '≈': '~',
            '≠': '!=',
            '≤': '<=',
            '≥': '>=',
            '°': ' degrees',
            'π': 'pi',
            'θ': 'theta',
            'α': 'alpha',
            'β': 'beta',
            'γ': 'gamma',
            'δ': 'delta',
            'λ': 'lambda',
            'μ': 'mu',
            'σ': 'sigma',
            'φ': 'phi',
            'ψ': 'psi',
            'ω': 'omega',
            '²': '^2',
            '³': '^3',
            '¹': '^1',
            '⁴': '^4',
            '⁵': '^5',
            '⁶': '^6',
            '⁷': '^7',
            '⁸': '^8',
            '⁹': '^9',
            '⁰': '^0',
            '×': 'x',
            '÷': '/',
            '±': '+/-',
            '→': '->',
            '←': '<-',
            '↔': '<->',
            '•': '*',
            '·': '.',
        }
        
        # Apply replacements
        for unicode_char, ascii_equiv in replacements.items():
            text = text.replace(unicode_char, ascii_equiv)
        
        # Try to convert remaining Unicode to ASCII
        try:
            # Normalize to decomposed form and remove combining characters
            text = unicodedata.normalize('NFKD', text)
            text = text.encode('ascii', 'ignore').decode('ascii')
        except:
            # If all else fails, remove non-ASCII characters
            text = re.sub(r'[^\x00-\x7F]+', '', text)
        
        return text
    
    @staticmethod
    def generate_excel(results: List[Dict[str, Any]]) -> bytes:
        """
        Generate Excel report with question-by-question analysis
        
        Args:
            results: List of analysis results for each question
            
        Returns:
            Excel file as bytes
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Question Analysis"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Mistake cell style
        mistake_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        correct_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        
        # Write headers
        headers = ["Question #", "Question", "Gemini's Solution", "Student's Answer", "Mistake/Evaluation"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for idx, result in enumerate(results, 2):
            has_mistake = result.get("has_mistake", False)
            
            # Apply conditional formatting based on mistake status
            row_fill = mistake_fill if has_mistake else correct_fill
            
            # Question Number
            cell = ws.cell(row=idx, column=1, value=result.get("question_number", idx-1))
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = border
            cell.fill = row_fill
            
            # Question Text
            cell = ws.cell(row=idx, column=2, value=result.get("question_text", "N/A"))
            cell.alignment = cell_alignment
            cell.border = border
            
            # Gemini's Solution
            cell = ws.cell(row=idx, column=3, value=result.get("gemini_solution", "N/A"))
            cell.alignment = cell_alignment
            cell.border = border
            
            # Student's Answer
            cell = ws.cell(row=idx, column=4, value=result.get("student_answer", "N/A"))
            cell.alignment = cell_alignment
            cell.border = border
            
            # Mistake/Evaluation
            evaluation_text = result.get("mistake_description", "N/A")
            if not has_mistake:
                evaluation_text = "CORRECT: " + evaluation_text
            else:
                evaluation_text = "MISTAKE: " + evaluation_text
            
            cell = ws.cell(row=idx, column=5, value=evaluation_text)
            cell.alignment = cell_alignment
            cell.border = border
            cell.fill = row_fill
            
            # Set row height based on content
            max_lines = max(
                len(str(result.get("question_text", "")).split('\n')),
                len(str(result.get("gemini_solution", "")).split('\n')),
                len(str(result.get("student_answer", "")).split('\n')),
                len(str(evaluation_text).split('\n'))
            )
            ws.row_dimensions[idx].height = max(60, max_lines * 15)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    @staticmethod
    def generate_pdf(results: List[Dict[str, Any]]) -> bytes:
        """
        Generate PDF report with question-by-question analysis
        
        Args:
            results: List of analysis results for each question
            
        Returns:
            PDF file as bytes
        """
        pdf = FPDF()
        pdf.set_margins(left=15, top=15, right=15)
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Title page
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 20, "Student Performance Analysis Report", ln=True, align="C")
        
        pdf.set_font("Helvetica", "", 12)
        pdf.cell(0, 10, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
        pdf.ln(10)
        
        # Summary statistics
        total_questions = len(results)
        questions_with_mistakes = sum(1 for r in results if r.get("has_mistake", False))
        questions_correct = total_questions - questions_with_mistakes
        
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Summary", ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 7, f"Total Questions: {total_questions}", ln=True)
        pdf.cell(0, 7, f"Correct Answers: {questions_correct}", ln=True)
        pdf.cell(0, 7, f"Mistakes Found: {questions_with_mistakes}", ln=True)
        pdf.ln(5)
        
        # Detailed analysis for each question
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Question-by-Question Analysis", ln=True)
        pdf.ln(5)
        
        for idx, result in enumerate(results, 1):
            question_num = result.get("question_number", idx)
            has_mistake = result.get("has_mistake", False)
            
            # Question header
            pdf.set_font("Helvetica", "B", 11)
            if has_mistake:
                pdf.set_fill_color(255, 230, 230)  # Light red for mistakes
                status = "MISTAKE FOUND"
            else:
                pdf.set_fill_color(230, 255, 230)  # Light green for correct
                status = "CORRECT"
            
            pdf.cell(0, 8, f"Question {question_num} - {status}", ln=True, fill=True)
            pdf.ln(2)
            
            # 1. Question Text
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, "Question:", ln=True)
            pdf.set_font("Helvetica", "", 9)
            question_text = ReportGenerator.sanitize_for_pdf(result.get("question_text", "N/A"))
            pdf.multi_cell(0, 5, question_text)
            pdf.ln(2)
            
            # 2. Gemini's Solution
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, "Gemini's Solution:", ln=True)
            pdf.set_font("Helvetica", "", 9)
            gemini_solution = ReportGenerator.sanitize_for_pdf(result.get("gemini_solution", "N/A"))
            pdf.multi_cell(0, 5, gemini_solution)
            pdf.ln(2)
            
            # 3. Student's Answer
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, "Student's Answer:", ln=True)
            pdf.set_font("Helvetica", "", 9)
            student_answer = ReportGenerator.sanitize_for_pdf(result.get("student_answer", "N/A"))
            pdf.multi_cell(0, 5, student_answer)
            pdf.ln(2)
            
            # 4. Evaluation/Mistake
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, "Evaluation:", ln=True)
            pdf.set_font("Helvetica", "", 9)
            mistake_desc = ReportGenerator.sanitize_for_pdf(result.get("mistake_description", "N/A"))
            pdf.multi_cell(0, 5, mistake_desc)
            
            pdf.ln(5)
            
            # Page break if needed
            if pdf.get_y() > 250:
                pdf.add_page()
        
        # Return PDF as bytes
        return bytes(pdf.output())
    
    @staticmethod
    def create_downloadable_excel(results: List[Dict[str, Any]], filename: str = "analysis_report.xlsx") -> tuple:
        """
        Create downloadable Excel file for Streamlit
        
        Args:
            results: Analysis results
            filename: Output filename
            
        Returns:
            Tuple of (file_bytes, filename)
        """
        excel_bytes = ReportGenerator.generate_excel(results)
        return excel_bytes, filename
    
    @staticmethod
    def create_downloadable_pdf(results: List[Dict[str, Any]], filename: str = "analysis_report.pdf") -> tuple:
        """
        Create downloadable PDF file for Streamlit
        
        Args:
            results: Analysis results
            filename: Output filename
            
        Returns:
            Tuple of (file_bytes, filename)
        """
        pdf_bytes = ReportGenerator.generate_pdf(results)
        return pdf_bytes, filename
